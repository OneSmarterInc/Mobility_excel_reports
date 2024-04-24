import sqlite3
import pandas as pd
import re
import psycopg2
import pdfplumber
filename = r"D:/project/Verizon Bills for Sagar/mob_1021_68577381200001_11202023.pdf"
conn = psycopg2.connect(database="mobilityprod", user="postgres", password="GetOutOfJail@2023", host="localhost", port="5432")
cursor = conn.cursor()
query = "SELECT * FROM tech_pdf_data_table"
dataframe = pd.read_sql_query(query, conn)
df = dataframe
df.columns = df.columns.str.replace('_', ' ').str.title()
df_unique = df.drop_duplicates(subset=['Wireless Number'])
print(df_unique.columns)
print("++++++++++++++++++++++_______________________")
table_1_data = df_unique
table_Bills_Details = df_unique
table_2_data = df_unique
unique_plans = table_2_data['Plans'].unique()
plan_discount_costs = {}
for plan in unique_plans:
    try:
        discount_data = table_2_data[(table_2_data['Plans'] == plan) &
                                    table_2_data['Item Description'].str.contains('Discount', na=False) &
                                    (table_2_data['Charges'] < 0)]
    except:
        discount_data = table_2_data[(table_2_data['Plans'] == plan) &
                                    table_2_data['Item Description'].str.contains('Discount', na=False) &
                                    (table_2_data['Charges'].str.replace('$','').astype(float) < 0)]

    discount_costs = discount_data['Charges'].unique()

    plan_discount_costs[plan] = discount_costs

unique_plans_df = pd.DataFrame(unique_plans, columns=['Unique Plans'])
discount_totals = {}

# Iterating over each unique plan
for plan in unique_plans:
    # Filtering data for the specific plan and rows with 'Discount' in "Item Description"
    plan_data = table_2_data[(table_2_data['Plans'] == plan) & table_2_data['Item Description'].str.contains('Discount', na=False)]

    # Calculating the total discount for the plan
    total_discount = plan_data['Charges'].astype(float).sum()

    # Storing the total discount in the dictionary
    discount_totals[plan] = total_discount

# Converting the dictionary to a DataFrame for better visualization
discount_totals_df = pd.DataFrame(list(discount_totals.items()), columns=['Plans', 'Total Discount'])
discount_data = table_2_data[table_2_data['Item Description'].str.contains('Discount', na=False)]

# Grouping the records by the Plans column and identifying if any line of service has a discount
grouped_discount_data = discount_data.groupby('Plans').apply(lambda group: group['Charges'].astype(float).sum())

# Getting the unique plans with discounts
plans_with_discounts = grouped_discount_data.index.tolist()

# For each plan with discounts, checking if all lines with that plan have the discount applied
missing_discounts = []
for plan in plans_with_discounts:
    total_lines_with_plan = table_2_data[table_2_data['Plans'] == plan]
    total_lines_with_discount = discount_data[discount_data['Plans'] == plan]

    if len(total_lines_with_plan) != len(total_lines_with_discount):
        missing_lines = total_lines_with_plan[~total_lines_with_plan['Wireless Number'].isin(total_lines_with_discount['Wireless Number'])]
        missing_discounts.extend(missing_lines[['User Name', 'Wireless Number', 'Plans']].to_dict('records'))

# Convert the results into a DataFrame
missing_discounts_df = pd.DataFrame(missing_discounts)

# Print the results
missing_discounts_df = missing_discounts_df.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
discount_data = table_2_data[table_2_data['Item Description'].str.contains('Discount', na=False)]

# Grouping the records by the Plans column and identifying if any line of service has a discount
grouped_discount_data = discount_data.groupby('Plans').apply(lambda group: group['Charges'].astype(float).sum())

# Getting the unique plans with discounts
plans_with_discounts = grouped_discount_data.index.tolist()

# For each plan with discounts, checking if all lines with that plan have the discount applied
missing_discounts = []
for plan in plans_with_discounts:
    total_lines_with_plan = table_2_data[table_2_data['Plans'] == plan]
    total_lines_with_discount = discount_data[discount_data['Plans'] == plan]

    if len(total_lines_with_plan) != len(total_lines_with_discount):
        missing_lines = total_lines_with_plan[~total_lines_with_plan['Wireless Number'].isin(total_lines_with_discount['Wireless Number'])]
        missing_discounts.extend(missing_lines[['User Name', 'Wireless Number', 'Plans']].to_dict('records'))

# Calculate the potential missed discount for each missing wireless number
for record in missing_discounts:
    plan = record['Plans']
    if plan in plans_with_discounts:
        average_discount = discount_data[discount_data['Plans'] == plan]['Charges'].astype(float).mean()
        record['Missed Discount'] = average_discount

# Convert the results into a DataFrame
missing_discounts_df = pd.DataFrame(missing_discounts)
missing_discounts_df = missing_discounts_df.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
data_device_criteria = table_2_data['Plans'].str.contains('Data Device|MIFI|Tablet|Data Dvc', case=False, na=False)
data_devices = table_2_data[data_device_criteria]

# Filtering basic based on plan names
Basic_criteria = table_2_data['Plans'].str.contains('Basic', case=False, na=False)
Basic = table_2_data[Basic_criteria]

# Filtering smartphones based on plan names
smartphone_criteria = table_2_data['Plans'].str.contains('Smartphone|Smartphn|Smrtphn|Phone', case=False, na=False)
smartphones = table_2_data[smartphone_criteria]
data_device_usage = table_1_data[table_1_data['Wireless Number'].isin(data_devices['Wireless Number'])][['Wireless Number', 'Data Usage']]
Basic_usage = table_1_data[table_1_data['Wireless Number'].isin(Basic['Wireless Number'])][['Wireless Number', 'Data Usage']]
smartphone_usage = table_1_data[table_1_data['Wireless Number'].isin(smartphones['Wireless Number'])][['Wireless Number', 'Data Usage']]
# Convert the 'Data Usage' column to numeric values by removing the 'GB' suffix
data_device_usage['Data Usage'] = data_device_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', 0).astype(float)
Basic_usage['Data Usage'] = Basic_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', 0).astype(float)
smartphone_usage['Data Usage'] = smartphone_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', 0).astype(float)

# Recalculate the average data usage for each line of service
data_device_usage['Average Data Usage'] = data_device_usage['Data Usage'].mean()
Basic_usage['Average Data Usage'] = Basic_usage['Data Usage'].mean()
smartphone_usage['Average Data Usage'] = smartphone_usage['Data Usage'].mean()

data_device_usage_sorted = data_device_usage[data_device_usage['Data Usage'] != 0].sort_values(by=['Data Usage'], ascending=False)

# For smartphones: First, we need to retrieve voice and text usage columns from the original data
smartphone_usage['Voice Usage'] = table_1_data['Voice Plan Usage']
smartphone_usage['Text Usage'] = table_1_data['Messaging Usage']

# Filter out lines with 0 data usage and custom sort using Data usage, Voice, and Text columns
# smartphone_usage_sorted = smartphone_usage[smartphone_usage['Data Usage'] != 0].sort_values(by=['Data Usage', 'Voice Usage', 'Text Usage'], ascending=False)
# print(temp_smart)
print('__________________))))))))))))))))))(((((((((((((((((((((((((1)))))))))))))))))))))))))')
try:
   smartphone_usage_sorted = smartphone_usage.reset_index()[smartphone_usage['Data Usage'] != 0].sort_values(by=['Data Usage', 'Voice Usage', 'Text Usage'], ascending=False)
except:
   smartphone_usage_sorted = smartphone_usage[smartphone_usage['Data Usage'] != 0].sort_values(by=['Data Usage', 'Voice Usage', 'Text Usage'], ascending=False)
data_device_zero_usage = data_device_usage[data_device_usage['Data Usage'] == 0]
smartphone_zero_usage = smartphone_usage[smartphone_usage['Data Usage'] == 0]
Basic_usage = Basic_usage[Basic_usage['Data Usage'] == 0]

# Remove the 0 usage devices from the sorted datasets
data_device_usage_sorted = data_device_usage_sorted[data_device_usage_sorted['Data Usage'] != 0]
smartphone_usage_sorted = smartphone_usage_sorted[smartphone_usage_sorted['Data Usage'] != 0]
data_device_usage_sorted = data_device_usage_sorted.sort_values(by='Average Data Usage', ascending=False)
smartphone_usage_sorted = smartphone_usage_sorted.sort_values(by='Average Data Usage', ascending=False)
plan_costs = []
print('__________________))))))))))))))))))(((((((((((((((((((((((((2)))))))))))))))))))))))))')
unique_plan = table_2_data["Plans"].unique()
print(unique_plan)
for plan in unique_plan:
        # print(plan)
        # item_list = list(table_2_data['Item Category'])
        # print(item_list)
        # temp_mon_data = table_2_data['Item Category'] == 'Monthly Charges'
        # print(temp_mon_data)
        # temp_df = table_2_data[temp_mon_data]
        if plan != 'NA':
          for index, row in df.iterrows():
            plan_costs.append({
                'Plans': plan,
                'Monthly Charge': row['Monthly Charges']
            })
print('__________________))))))))))))))))))(((((((((((((((((((((((((3)))))))))))))))))))))))))')
# Convert the results to a DataFrame for display
plan_costs_df = pd.DataFrame(plan_costs)
print('__________________))))))))))))))))))(((((((((((((((((((((((((4)))))))))))))))))))))))))')
plan_costs_df = plan_costs_df.drop_duplicates(subset=['Plans'], ignore_index=True)
plan_costs_df['Monthly Charge'] = plan_costs_df['Monthly Charge'].str.replace('$','').astype(float)
data_usage = table_1_data[['Wireless Number', 'Data Usage']]
data_usage['Data Usage'] = data_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', '0').astype(float)
user_data_plan = pd.merge(data_usage, table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='inner')
user_data_plan = user_data_plan.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
user_data_plan = pd.merge(user_data_plan, plan_costs_df[['Plans', 'Monthly Charge']], on='Plans', how='left')
user_data_plan['Data Usage'] = user_data_plan['Data Usage'].astype(float)
user_data_plan['Monthly Charge'] = user_data_plan['Monthly Charge'].astype(float)
def extract_data_usage(plan):
    match = re.search(r'(\d+\.?\d*)(GB|MB)', plan)
    if match:
        value = float(match.group(1))
        if match.group(2) == "MB":
            value = value / 1000
        return value
    return None

plan_costs_df['Data Usage (GB)'] = plan_costs_df['Plans'].apply(extract_data_usage)

def find_best_plan(row):
    if row['Data Usage'] == 0 or row['Plans'] == 'NA':
        return row['Plans']
    current_plan_data = plan_costs_df[plan_costs_df['Plans'] == row['Plans']].iloc[0]
    potential_plans = plan_costs_df[
        (plan_costs_df['Plans'].str.contains(re.split(r'\d+GB|\d+MB', current_plan_data['Plans'])[0])) &
        (plan_costs_df['Data Usage (GB)'] < current_plan_data['Data Usage (GB)'])
    ].sort_values(by='Data Usage (GB)', ascending=False)
    if potential_plans.empty:
        return row['Plans']
    for _, plan in potential_plans.iterrows():
        if plan['Data Usage (GB)'] >= row['Data Usage']:
            return plan['Plans']
    return row['Plans']

user_data_plan['Best Plan'] = user_data_plan.apply(find_best_plan, axis=1)
user_data_plan = pd.merge(user_data_plan, plan_costs_df[['Plans', 'Monthly Charge']], left_on='Best Plan', right_on='Plans', how='left')
user_data_plan = user_data_plan.rename(columns={'Plans_x': 'Current Plan', 'Plans_y': 'Best Plan Name'})
user_data_plan['Monthly Charge Difference'] = user_data_plan['Monthly Charge_x'] - user_data_plan['Monthly Charge_y']

# Ensure the columns being subtracted contain only numeric values
user_data_plan['Monthly Charge_x'] = pd.to_numeric(user_data_plan['Monthly Charge_x'], errors='coerce')
user_data_plan['Monthly Charge_y'] = pd.to_numeric(user_data_plan['Monthly Charge_y'], errors='coerce')

# Displaying the final dataframe
user_data_plan = user_data_plan[['Wireless Number', 'Data Usage', 'Current Plan', 'Best Plan', 'Monthly Charge_x', 'Monthly Charge_y', 'Monthly Charge Difference']]
user_data_plan
def extract_data_allowance(plan_name):
    if "GB" in plan_name:
        return float(''.join(filter(str.isdigit, plan_name.split("GB")[0])))
    elif "MB" in plan_name:
        return round(float(''.join(filter(str.isdigit, plan_name.split("MB")[0]))) / 1024, 3)  # Convert MB to GB
    else:
        return "Unlimited"

# Function to get total data usage for a plan
def get_total_data_usage(plan):
    Wireless_Numbers = table_2_data[table_2_data['Plans'] == plan]['Wireless Number'].unique()
    data_usage = table_1_data[table_1_data['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
    total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)GB').dropna().astype(float).sum()
    return total_usage[0]

# Extract unique plan names excluding "NA"
unique_plans = [plan for plan in table_2_data['Plans'].unique() if plan != "NA"]

# Create the dataframe with the modifications
plans_data_allowances = pd.DataFrame({
    'Plan Name': unique_plans,
    'Data Allowance': [extract_data_allowance(plan) for plan in unique_plans]
})

# Populate the dataframe with the required data
plans_data_allowances['Qty of Wireless Numbers'] = plans_data_allowances['Plan Name'].apply(
    lambda plan: table_2_data[table_2_data['Plans'] == plan]['Wireless Number'].nunique())
plans_data_allowances['Total Data Allowance'] = plans_data_allowances.apply(
    lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
plans_data_allowances['Total Data Usage'] = plans_data_allowances['Plan Name'].apply(get_total_data_usage)

# Create separate DataFrames based on the keywords
def categorize_plan(plan_name):
    smartphone_keywords = ['Smartphone', 'Smartphn', 'Smrtphn']
    data_device_keywords = ['Data Device', 'Data Dvc']

    if any(keyword in plan_name for keyword in smartphone_keywords):
        return "smartphone"
    elif any(keyword in plan_name for keyword in data_device_keywords):
        return "data_device"
    else:
        return "other"

# Apply categorize_plan function to categorize plans into smartphone, data_device, and other
plans_data_allowances['Category'] = plans_data_allowances['Plan Name'].apply(categorize_plan)

# Calculate the totals
total_data_allowance = plans_data_allowances[plans_data_allowances['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage = plans_data_allowances['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance],
    'Total Data Usage': [total_data_usage],
    'Category': ['NA']
})

plans_data_allowances = pd.concat([plans_data_allowances, totals_row], ignore_index=True)

# Display the result
plans_data_allowances
smartphone_df = plans_data_allowances[plans_data_allowances['Category'] == "smartphone"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance = smartphone_df[smartphone_df['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage = smartphone_df['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance],
    'Total Data Usage': [total_data_usage],
})

smartphone_df = smartphone_df.append(totals_row, ignore_index=True)
data_device_df = plans_data_allowances[plans_data_allowances['Category'] == "data_device"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance = data_device_df[data_device_df['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage = data_device_df['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance],
    'Total Data Usage': [total_data_usage],
})

data_device_df = data_device_df.append(totals_row, ignore_index=True)

other_df = plans_data_allowances[plans_data_allowances['Category'] == "other"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance = other_df[other_df['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage = other_df['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance],
    'Total Data Usage': [total_data_usage],
})

other_df = other_df.append(totals_row, ignore_index=True)
ptrt = table_1_data['Data Usage'].str.extract(r'(.\d+|\d+.\d+)GB').astype(float)[0] >= 15.0
high_data_usage = table_1_data[ptrt]
# Merge with table_2_data to get the Plans column
high_data_usage = high_data_usage.merge(table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='left')
# Reset the index
high_data_usage = high_data_usage.reset_index(drop=True)
# Select the desired columns
if not high_data_usage.empty:
    result_high_data_usage = high_data_usage[['Wireless Number', 'User Name', 'Monthly Charges', 'Plans_y', 'Data Usage']]
    result_high_data_usage = result_high_data_usage.drop_duplicates(subset=['Wireless Number'], ignore_index=True)
zero_data_usage = table_1_data[table_1_data['Data Usage'].str.extract(r'(NA)').replace('NA', '0').astype(float)[0] >= 0.0]
zero_data_usage = zero_data_usage.merge(table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='left')
zero_data_usage = zero_data_usage.reset_index(drop=True)
result_zero_data_usage = zero_data_usage[['Wireless Number', 'User Name', 'Monthly Charges', 'Plans_y', 'Data Usage']]
result_zero_data_usage = result_zero_data_usage.drop_duplicates(subset=['Wireless Number'], ignore_index=True)
# Modify the original table_1_data to exclude wireless numbers with data usage >= 15.00 GB
lower_data_usage = table_1_data[~table_1_data['Wireless Number'].isin(high_data_usage['Wireless Number'])]
lower_data_usage = lower_data_usage[~lower_data_usage['Wireless Number'].isin(zero_data_usage['Wireless Number'])]

# Merge with table_2_data to get the Plans column
lower_data_usage = lower_data_usage.merge(table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='left')

# Reset the index
lower_data_usage = lower_data_usage.reset_index(drop=True)

# Select the desired columns
result_lower_data_usage = lower_data_usage[['Wireless Number', 'User Name', 'Monthly Charges', 'Plans_y', 'Data Usage']]
result_lower_data_usage = result_lower_data_usage.drop_duplicates(subset=['Wireless Number'], ignore_index=True)
filtered_table_1_data = table_1_data[~table_1_data['Wireless Number'].isin(high_data_usage['Wireless Number'])]
filtered_table_1_data = filtered_table_1_data[~filtered_table_1_data['Wireless Number'].isin(zero_data_usage['Wireless Number'])]
filtered_table_1_data = filtered_table_1_data.reset_index(drop=True)
filtered_table_2_data = table_2_data[~table_2_data['Wireless Number'].isin(high_data_usage['Wireless Number'])]
filtered_table_2_data = filtered_table_2_data[~filtered_table_2_data['Wireless Number'].isin(zero_data_usage['Wireless Number'])]
filtered_table_2_data = filtered_table_2_data.reset_index(drop=True)
# Function to extract data allowance from plan name
def extract_data_allowance(plan_name):
    if "GB" in plan_name:
        return float(''.join(filter(str.isdigit, plan_name.split("GB")[0])))
    elif "MB" in plan_name:
        return round(float(''.join(filter(str.isdigit, plan_name.split("MB")[0]))) / 1024, 3)  # Convert MB to GB
    else:
        return "Unlimited"

# Function to get total data usage for a plan
def get_total_data_usage(plan):
    Wireless_Numbers = filtered_table_2_data[filtered_table_2_data['Plans'] == plan]['Wireless Number'].unique()
    data_usage = filtered_table_1_data[filtered_table_1_data['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
    total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)GB').dropna().astype(float).sum()
    return total_usage[0]

# Extract unique plan names excluding "NA"
unique_plans = [plan for plan in filtered_table_2_data['Plans'].unique() if plan != "NA"]

# Create the dataframe with the modifications
plans_data_allowances1 = pd.DataFrame({
    'Plan Name': unique_plans,
    'Data Allowance': [extract_data_allowance(plan) for plan in unique_plans]
})

# Populate the dataframe with the required data
plans_data_allowances1['Qty of Wireless Numbers'] = plans_data_allowances1['Plan Name'].apply(
    lambda plan: filtered_table_2_data[filtered_table_2_data['Plans'] == plan]['Wireless Number'].nunique())
plans_data_allowances1['Total Data Allowance'] = plans_data_allowances1.apply(
    lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
plans_data_allowances1['Total Data Usage'] = plans_data_allowances1['Plan Name'].apply(get_total_data_usage)

# Create separate DataFrames based on the keywords
def categorize_plan(plan_name):
    smartphone_keywords = ['Smartphone', 'Smartphn', 'Smrtphn']
    data_device_keywords = ['Data Device', 'Data Dvc']

    if any(keyword in plan_name for keyword in smartphone_keywords):
        return "smartphone"
    elif any(keyword in plan_name for keyword in data_device_keywords):
        return "data_device"
    else:
        return "other"

# Apply categorize_plan function to categorize plans into smartphone, data_device, and other
plans_data_allowances1['Category'] = plans_data_allowances1['Plan Name'].apply(categorize_plan)

# Calculate the totals
total_data_allowance1 = plans_data_allowances1[plans_data_allowances1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage1 = plans_data_allowances1['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row1 = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance1],
    'Total Data Usage': [total_data_usage1],
    'Category': ['NA']
})

plans_data_allowances1 = plans_data_allowances1.append(totals_row1, ignore_index=True)
smartphone_df1 = plans_data_allowances1[plans_data_allowances1['Category'] == "smartphone"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance1 = smartphone_df1[smartphone_df1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage1 = smartphone_df1['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance1],
    'Total Data Usage': [total_data_usage1],
})

smartphone_df1 = smartphone_df1.append(totals_row, ignore_index=True)
data_device_df1 = plans_data_allowances1[plans_data_allowances1['Category'] == "data_device"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance1 = data_device_df1[data_device_df1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage1 = data_device_df1['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance1],
    'Total Data Usage': [total_data_usage1],
})

data_device_df1 = data_device_df1.append(totals_row, ignore_index=True)
other_df1 = plans_data_allowances1[plans_data_allowances1['Category'] == "other"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance1 = other_df1[other_df1['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage1 = other_df1['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance1],
    'Total Data Usage': [total_data_usage1],
})

other_df1 = other_df1.append(totals_row, ignore_index=True)
plan_costs = []
unique_plan = filtered_table_2_data["Plans"].unique()

for plan in unique_plans:
        # Filter rows where the 'Item Category' is 'Monthly Charges'
        if plan != 'NA':
          for index, row in df.iterrows():
            plan_costs.append({
                'Plans': plan,
                'Monthly Charge': row['Charges']
            })

# Convert the results to a DataFrame for display
plan_costs_df1 = pd.DataFrame(plan_costs)
plan_costs_df1 = plan_costs_df1.drop_duplicates(subset=['Plans'], ignore_index=True)
plan_costs_df1['Monthly Charge'] = plan_costs_df1['Monthly Charge'].str.replace('$','').astype(float)
data_usage = filtered_table_1_data[['Wireless Number', 'Data Usage']]
data_usage.loc[:, 'Data Usage'] = data_usage['Data Usage'].str.replace('GB', '', regex=False).replace('NA', '0').astype(float)
user_data_plan1 = pd.merge(data_usage, filtered_table_2_data[['Wireless Number', 'Plans']], on='Wireless Number', how='inner')
user_data_plan1 = user_data_plan1.drop_duplicates(subset=['Wireless Number', 'Plans'], ignore_index=True)
user_data_plan1 = pd.merge(user_data_plan1, plan_costs_df1[['Plans', 'Monthly Charge']], on='Plans', how='left')
user_data_plan1['Data Usage'] = user_data_plan1['Data Usage'].astype(float)
user_data_plan1['Monthly Charge'] = user_data_plan1['Monthly Charge'].astype(float)
def extract_data_usage(plan):
    match = re.search(r'(\d+\.?\d*)(GB|MB)', plan)
    if match:
        value = float(match.group(1))
        if match.group(2) == "MB":
            value = value / 1000
        return value
    return None

plan_costs_df1['Data Usage (GB)'] = plan_costs_df1['Plans'].apply(extract_data_usage)

def find_best_plan(row):
    if row['Data Usage'] == 0 or row['Plans'] == 'NA':
        return row['Plans']
    current_plan_data = plan_costs_df1[plan_costs_df1['Plans'] == row['Plans']].iloc[0]
    potential_plans = plan_costs_df1[
        (plan_costs_df1['Plans'].str.contains(re.split(r'\d+GB|\d+MB', current_plan_data['Plans'])[0])) &
        (plan_costs_df1['Data Usage (GB)'] < current_plan_data['Data Usage (GB)'])
    ].sort_values(by='Data Usage (GB)', ascending=False)
    if potential_plans.empty:
        return row['Plans']
    for _, plan in potential_plans.iterrows():
        if plan['Data Usage (GB)'] >= row['Data Usage']:
            return plan['Plans']
    return row['Plans']

user_data_plan1['Best Plan'] = user_data_plan1.apply(find_best_plan, axis=1)
user_data_plan1 = pd.merge(user_data_plan1, plan_costs_df1[['Plans', 'Monthly Charge']], left_on='Best Plan', right_on='Plans', how='left')
user_data_plan1 = user_data_plan1.rename(columns={'Plans_x': 'Current Plan', 'Plans_y': 'Best Plan Name'})
user_data_plan1['Monthly Charge Difference'] = user_data_plan1['Monthly Charge_x'] - user_data_plan1['Monthly Charge_y']

# Ensure the columns being subtracted contain only numeric values
user_data_plan1['Monthly Charge_x'] = pd.to_numeric(user_data_plan1['Monthly Charge_x'], errors='coerce')
user_data_plan1['Monthly Charge_y'] = pd.to_numeric(user_data_plan1['Monthly Charge_y'], errors='coerce')

# Displaying the final dataframe
user_data_plan1 = user_data_plan1[['Wireless Number', 'Data Usage', 'Current Plan', 'Best Plan', 'Monthly Charge_x', 'Monthly Charge_y', 'Monthly Charge Difference']]
def extract_data_allowance(plan_name):
    if "GB" in plan_name:
        return float(''.join(filter(str.isdigit, plan_name.split("GB")[0])))
    elif "MB" in plan_name:
        return round(float(''.join(filter(str.isdigit, plan_name.split("MB")[0]))) / 1024, 3)  # Convert MB to GB
    else:
        return "Unlimited"

# Function to get total data usage for a plan
def get_total_data_usage(plan):
    Wireless_Numbers = user_data_plan1[user_data_plan1['Best Plan'] == plan]['Wireless Number'].unique()
    data_usage = user_data_plan1[user_data_plan1['Wireless Number'].isin(Wireless_Numbers)]['Data Usage']
    data_usage = data_usage.astype(str)
    total_usage = data_usage.str.extract(r'(.\d+|\d+.\d+)').dropna().astype(float).sum()
    return total_usage[0]

# Extract unique plan names excluding "NA"
unique_plans = [plan for plan in user_data_plan1['Best Plan'].unique() if plan != "NA"]

# Create the dataframe with the modifications
plans_data_allowances2 = pd.DataFrame({
    'Plan Name': unique_plans,
    'Data Allowance': [extract_data_allowance(plan) for plan in unique_plans]
})

# Populate the dataframe with the required data
plans_data_allowances2['Qty of Wireless Numbers'] = plans_data_allowances2['Plan Name'].apply(
    lambda plan: user_data_plan1[user_data_plan1['Best Plan'] == plan]['Wireless Number'].nunique())
plans_data_allowances2['Total Data Allowance'] = plans_data_allowances2.apply(
    lambda row: row['Data Allowance'] * row['Qty of Wireless Numbers'] if row['Data Allowance'] != "Unlimited" else "Unlimited", axis=1)
plans_data_allowances2['Total Data Usage'] = plans_data_allowances2['Plan Name'].apply(get_total_data_usage)

# Create separate DataFrames based on the keywords
def categorize_plan(plan_name):
    smartphone_keywords = ['Smartphone', 'Smartphn', 'Smrtphn']
    data_device_keywords = ['Data Device', 'Data Dvc']

    if any(keyword in plan_name for keyword in smartphone_keywords):
        return "smartphone"
    elif any(keyword in plan_name for keyword in data_device_keywords):
        return "data_device"
    else:
        return "other"

# Apply categorize_plan function to categorize plans into smartphone, data_device, and other
plans_data_allowances2['Category'] = plans_data_allowances2['Plan Name'].apply(categorize_plan)

# Calculate the totals
total_data_allowance2 = plans_data_allowances2[plans_data_allowances2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage2 = plans_data_allowances2['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row1 = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance2],
    'Total Data Usage': [total_data_usage2],
    'Category': ['NA']
})

plans_data_allowances2 = pd.concat([plans_data_allowances2, totals_row1], ignore_index=True)
smartphone_df2 = plans_data_allowances2[plans_data_allowances2['Category'] == "smartphone"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance2 = smartphone_df2[smartphone_df2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage2 = smartphone_df2['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance2],
    'Total Data Usage': [total_data_usage2],
})

smartphone_df2 = pd.concat([smartphone_df2, totals_row], ignore_index=True)
data_device_df2 = plans_data_allowances2[plans_data_allowances2['Category'] == "data_device"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance2 = data_device_df2[data_device_df2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage2 = data_device_df2['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance2],
    'Total Data Usage': [total_data_usage2],
})

data_device_df2 = pd.concat([data_device_df2, totals_row], ignore_index=True)
other_df2 = plans_data_allowances2[plans_data_allowances2['Category'] == "other"].drop(columns=['Category'])

# Calculate the totals
total_data_allowance2 = other_df2[other_df2['Total Data Allowance'] != "Unlimited"]['Total Data Allowance'].sum()
total_data_usage2 = other_df2['Total Data Usage'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Plan Name': ['Total'],
    'Data Allowance': ['NA'],
    'Qty of Wireless Numbers': ['NA'],
    'Total Data Allowance': [total_data_allowance2],
    'Total Data Usage': [total_data_usage2],
})

other_df2 = pd.concat([other_df2, totals_row], ignore_index=True)
total_cost_current = user_data_plan1['Monthly Charge_x'].sum()
total_cost_best = user_data_plan1['Monthly Charge_y'].sum()
total_cost_save = user_data_plan1['Monthly Charge Difference'].sum()

# Append the totals row to the dataframe
totals_row = pd.DataFrame({
    'Wireless Number': ['Total Charges'],
    'Data Usage': ['NA'],
    'Current Plan': ['NA'],
    'Best Plan': ['NA'],
    'Monthly Charge_x': [total_cost_current],
    'Monthly Charge_y': [total_cost_best],
    'Monthly Charge Difference': [total_cost_save]
})

user_data_plan2 = pd.concat([user_data_plan1, totals_row], ignore_index=True)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Create a new Excel workbook
wb = Workbook()

# Create worksheets
ws1 = wb.create_sheet(title='15GB+ Data Usage')
ws2 = wb.create_sheet(title='0 Data Usage')
ws3 = wb.create_sheet(title='Except 15GB+ && 0 Data Usage')
ws4 = wb.create_sheet(title='Recomendations using user data')
ws5 = wb.create_sheet(title='All in One Bucket Analysis')
ws6 = wb.create_sheet(title='Smartphone Bucket Analysis')
ws7 = wb.create_sheet(title='Data Device Bucket Analysis')

# Write DataFrames to respective sheets
dataframes = {
    '15GB+ Data Usage': result_high_data_usage if 'result_high_data_usage' in locals() else 'NA',
    '0 Data Usage': result_zero_data_usage,
    'Except 15GB+ && 0 Data Usage': result_lower_data_usage,
    'Recomendations using user data':user_data_plan2,
    'All in One Bucket Analysis': plans_data_allowances2,
    'Smartphone Bucket Analysis': smartphone_df2,
    'Data Device Bucket Analysis': data_device_df2
}
total_sheets = []
dataframes = {key: value for key, value in dataframes.items() if not isinstance(value, str) or value != 'NA'}
for sheet_name, df in dataframes.items():
    sheet = wb[sheet_name]
    total_sheets.append(sheet)
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

import os
directory = 'op_test_dir'
if not os.path.exists(directory):
    os.makedirs(directory)

excel_files = total_sheets[:-3]
recommendations_worksheet = excel_files[-1]
recommendations_df = pd.DataFrame(recommendations_worksheet.values)
recommendations_header = recommendations_df.iloc[0]
recommendations_df = recommendations_df[1:]
recommendations_df.columns = recommendations_header

# Initialize the merged DataFrame with the data from the "Recommendations Using User Data Usage" worksheet
merged_df = recommendations_df.copy()

# List of other worksheet objects
other_worksheets = excel_files[:-1]
plan_dict = {
    "Business Unlimited Smartphone":1000,
    "Flex Business Data Device 2GB":2,
    "Business UNL Plus Smartphone":1000,
    "Business UNL Tablet Start":1000,
    "BUS UNL Plus 2.0 Smartphone":1000,
    "Business Unlimited Data":1000,
    "The new Verizon Plan BUS 25GB":25,
    "SMB UNL TLK&TXT 200GB":200,
    "More Evr SMB UNL TLK&TXT 400GB":400
}
pricing_dictionary = {
    "Business Unlimited Smartphone": 45,
    "Flex Business Data Device 2GB": 31.5,
    "Business UNL Plus Smartphone": 40,
    "Business UNL Tablet Start": 30,
    "BUS UNL Plus 2.0 Smartphone": 35,
    "Business Unlimited Data": 45,
    "The new Verizon Plan BUS 25GB": 136,
    "SMB UNL TLK&TXT 200GB": 1140,
    "More Evr SMB UNL TLK&TXT 400GB": 1170
}
# Iterate over each of the other worksheets
for worksheet in other_worksheets:
    # Read the worksheet into a DataFrame
    df = pd.DataFrame(worksheet.values)
    header = df.iloc[0]
    df = df[1:]
    df.columns = header
    
    # Iterate over each row in the current DataFrame
    for index, row in df.iterrows():
        wireless_number = row['Wireless Number']
        
        # Find the corresponding row in the "Recommendations Using User Data Usage" DataFrame based on the wireless number
        matching_row = recommendations_df[recommendations_df['Wireless Number'] == wireless_number]
        
        # If a matching row is found, merge the data from the current DataFrame into that row in the merged DataFrame
        if not matching_row.empty:
            for col in df.columns:
                merged_df.at[matching_row.index[0], col] = row[col]
        else:
            # If no matching row is found, append the row from the current DataFrame to the merged DataFrame
            merged_df = merged_df.append(row, ignore_index=True)

merged_df['Current Plan'] = merged_df.apply(lambda row: row['Plans_y'] if pd.isna(row['Current Plan']) else row['Current Plan'], axis=1)
def best_recommended_plan(data_usage):
    if data_usage == 'NA':
        return 'NA'
    data_usage = str(data_usage)
    print(type(data_usage))
    data_usage = data_usage.strip()
    if 'GB' in data_usage:
        data_usage = data_usage.replace('GB', '')
    
    try:
        data_usage = float(data_usage)
    except:
        data_usage = 0

    best_plan = None
    best_value = float('inf')
    
    data_usage = float(data_usage)
    
    for plan, value in plan_dict.items():
        if value > data_usage and value < best_value:
            best_plan = plan
            best_value = value
    
    return best_plan

merged_df['Best Plan'] = merged_df['Data Usage'].apply(best_recommended_plan)
plan_value_dict = {
    "Current Plan":"Monthly Charge_x",
    "Best Plan":"Monthly Charge_y"
}
for k in plan_value_dict:
    for index, row in merged_df.iterrows():
        current_plan = row[k]
        if current_plan in pricing_dictionary:
            merged_df.at[index, plan_value_dict[k]] = pricing_dictionary[current_plan]
merged_df['Recommendations'] = ""

for index, row in merged_df.iterrows():
    current_plan = row['Current Plan']
    best_plan = row['Best Plan']

    if best_plan == 'NA':
        merged_df.at[index, 'Recommendations'] = "Verify 0 usage-Cancel"
    elif current_plan == best_plan:
        merged_df.at[index, 'Recommendations'] = ""
    else:
        merged_df.at[index, 'Recommendations'] = "Change Plan"

merged_df = merged_df.rename(columns={'Wireless Number': 'Mobile no',
                                      'Best Plan': 'Recommended plan',
                                      'Plans_y': 'Plans',
                                      'Monthly Charge_x': 'Current $',
                                      'Monthly Charge_y': 'Recommended $'})

# Reorder columns according to the specified order
merged_df = merged_df[['Recommendations', 'User Name', 'Mobile no', 'Current Plan', 'Current $',
                       'Recommended plan', 'Recommended $', 'Plans', 'Data Usage']]
merged_df['Charge Difference'] = merged_df['Current $'] - merged_df['Recommended $']
# Path to the merged Excel file in the new directory
w8 = wb.create_sheet(title='Recomendations As Per User Data')
dataframess = {'Recomendations As Per User Data':merged_df}
dataframes = {key: value for key, value in dataframess.items() if not isinstance(value, str) or value != 'NA'}
for sheet_name, df in dataframes.items():
    sheet = wb[sheet_name]
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
wb.remove(ws4)
wb.move_sheet(w8, offset=-3)
excel_file_path = 'Verizon Arch bill PDF no 8.xlsx'
wb.save(excel_file_path)

